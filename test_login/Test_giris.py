from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager . chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gc
import openpyxl
import pytest
from selenium.webdriver.common.alert import Alert

class Test_giris_kontrol:   
        
        def getInvalidData():
                giris_excel = openpyxl.load_workbook("data/giris.xlsx")
                sheet = giris_excel["basarisiz"]
                rows = sheet.max_row
                data = []
                for i in range(2,rows+1):
                  eposta = sheet.cell(i,1).value
                  sifre = sheet.cell(i,2).value
                  data.append((eposta,sifre))
                return data
        def getValidData():
                giris_excel = openpyxl.load_workbook("data/giris.xlsx")
                sheet = giris_excel["basarili"]
                rows = sheet.max_row
                data = []
                for i in range(2,rows+1):
                  eposta = sheet.cell(i,1).value
                  sifre = sheet.cell(i,2).value
                  data.append((eposta,sifre))
                return data
        

        def  giris_ekrani(self):
            self.driver = webdriver.Chrome()
            self.driver.get("https://tobeto.com/giris")
            self.driver.maximize_window() 
            
        @pytest.mark.parametrize("eposta,sifre" , getValidData())    
        def test_basarili_giris(self,eposta,sifre):
            self.giris_ekrani()
            email = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME, gc.NAME_TAG)))
            email.send_keys(eposta)
            sleep(3)
            psw = WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.NAME, gc.PSW_TAG)))
            psw.send_keys(sifre)
            loginbutton = self.driver.find_element(By.XPATH, gc.GIRIS_YAP_XPATH)
            loginbutton.click()
            sleep(4)
            
        @pytest.mark.parametrize("eposta,sifre" , getInvalidData())
        def test_basarisiz_giris(self,eposta,sifre):
            self.giris_ekrani()
            email = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME, gc.NAME_TAG)))
            email.send_keys(eposta)
            psw = WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.NAME, gc.PSW_TAG)))
            psw.send_keys(sifre)
            sleep(2)
            loginbutton = self.driver.find_element(By.XPATH, gc.GIRIS_YAP_XPATH)
            loginbutton.click()
            try:
                alert_element = WebDriverWait(self.driver, 10).until(
                ec.presence_of_element_located((By.CLASS_NAME, "toast-body")))
                
                alert_message = alert_element.text
                print("Başarisiz Giriş Uyarisi:", alert_message)

            except Exception as e:
                print("Başarisiz giriş uyarisi beklenirken hata oluştu veya uyari bulunamadi:", str(e))
            sleep(5)
        
        def test_empty_login(self):
            self.giris_ekrani()
            email = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME, gc.NAME_TAG)))
            email.send_keys("@gmail.com")
            sleep(3)
            psw = WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.NAME, gc.PSW_TAG)))
            psw.send_keys("")
            loginbutton = self.driver.find_element(By.XPATH, gc.GIRIS_YAP_XPATH)
            loginbutton.click()
            empty_warning = self.driver.find_element(By.XPATH, "//*[@id='__next']/div/main/section/div/div/div[1]/div/form/p")
            empty_warning.text == "Doldurulması zorunlu alan*"
            sleep(4)
            
                
